# live_feed.py
# Real-time Nigerian plate recognition from camera/RTSP/video feed
# Logs detections to live_log.xlsx

import cv2
import numpy as np
import re
import os
import time
import threading
from datetime import datetime
from collections import Counter, deque
from dataclasses import dataclass, field
from typing import Optional, List, Tuple

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ── Import from your existing backend ────────────────────────
# We import the processing functions directly (no HTTP needed)
# This runs the models in the SAME process = faster

import sys
BACKEND_DIR = os.path.dirname(os.path.abspath(__file__))
if BACKEND_DIR not in sys.path:
    sys.path.insert(0, BACKEND_DIR)

from ultralytics import YOLO
from paddleocr import PaddleOCR

from statelga import (
    enrich_plate_data,
    validate_plate_format,
    get_state_from_plate,
)

# ══════════════════════════════════════════════════════════════
# CONFIGURATION
# ══════════════════════════════════════════════════════════════

CONFIG = {
    # ── Source ────────────────────────────────────────────────
    # Use 0 for built-in webcam
    # Use 1 for USB camera
    # Use "rtsp://..." for IP camera
    # Use "path/to/video.mp4" for file
    "source"            : 0,

    # ── Model paths ──────────────────────────────────────────
    "yolo_model"        : "best.pt",

    # ── Processing ───────────────────────────────────────────
    "process_every_n"   : 3,          # Process every Nth frame
    "imgsz"             : 640,        # YOLO input size
    "min_confidence"    : 40,         # Min OCR confidence %
    "min_yolo_conf"     : 0.5,        # Min YOLO detection confidence
    "crop_pad"          : 12,         # Padding around detected plate

    # ── Tracking ─────────────────────────────────────────────
    "track_timeout"     : 5.0,        # Seconds before a tracked plate expires
    "min_seen_count"    : 2,          # Must see plate this many times to confirm
    "similarity_thresh" : 0.6,        # Character similarity threshold for grouping
    "vote_agreement"    : 0.5,        # Character voting agreement threshold

    # ── Display ──────────────────────────────────────────────
    "show_window"       : True,       # Show live CV2 window
    "window_width"      : 1280,       # Display window width
    "window_height"     : 720,        # Display window height
    "box_color"         : (0, 255, 0),   # Green bounding box
    "box_thickness"     : 2,
    "text_color"        : (255, 255, 255),
    "text_bg_color"     : (0, 100, 0),

    # ── Logging ──────────────────────────────────────────────
    "excel_log"         : "live_log.xlsx",
    "max_log_entries"   : 10000,
    "auto_save_interval": 30,         # Save Excel every N seconds
}

# ══════════════════════════════════════════════════════════════
# PLATE FORMAT CONSTANTS
# ══════════════════════════════════════════════════════════════

STRICT_PLATE = re.compile(r'^[A-Z]{3}\d{3}[A-Z]{2}$')

NOISE_WORDS = [
    "FEDERAL", "REPUBLIC", "NIGERIA", "CENTRE", "CENTER",
    "UNITY", "EXCELLENCE", "STATE", "GOVERNMENT", "OF",
    "LAGOS", "ABUJA", "FCT",
]

DIGIT_TO_LETTER = {
    '0':'O','1':'I','2':'Z','3':'B','4':'A',
    '5':'S','6':'G','7':'T','8':'B','9':'D',
}

LETTER_TO_DIGIT = {
    'O':'0','I':'1','L':'1','Z':'2','B':'8',
    'S':'5','G':'6','T':'7','A':'4','E':'3',
    'D':'0','Q':'0','J':'7','C':'0','F':'7',
}

# ══════════════════════════════════════════════════════════════
# TRACKED PLATE DATACLASS
# ══════════════════════════════════════════════════════════════

@dataclass
class TrackedPlate:
    """Represents a plate being tracked across frames."""
    readings     : List[str]   = field(default_factory=list)
    confidences  : List[float] = field(default_factory=list)
    first_seen   : float       = 0.0
    last_seen    : float       = 0.0
    bbox         : tuple       = (0, 0, 0, 0)
    confirmed    : Optional[str] = None
    conf_score   : float       = 0.0
    state        : str         = "Unknown"
    logged       : bool        = False

    @property
    def seen_count(self) -> int:
        return len(self.readings)

    def add_reading(self, plate: str, conf: float, bbox: tuple):
        self.readings.append(plate)
        self.confidences.append(conf)
        self.last_seen = time.time()
        self.bbox = bbox
        if not self.first_seen:
            self.first_seen = self.last_seen

# ══════════════════════════════════════════════════════════════
# EXCEL LOGGER
# ══════════════════════════════════════════════════════════════

class LiveExcelLogger:
    """Logs confirmed plates to live_log.xlsx."""

    HEADERS = [
        "#", "Plate Number", "State",
        "Confidence (%)", "Seen Count",
        "First Seen", "Last Seen",
        "Duration (s)", "Timestamp",
    ]

    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    VALID_FILL  = PatternFill("solid", fgColor="E2EFDA")
    ALT_FILL    = PatternFill("solid", fgColor="F5F5F5")

    def __init__(self, path: str):
        self.path   = path
        self._lock  = threading.Lock()
        self._wb    = openpyxl.Workbook()
        self._ws    = self._wb.active
        self._ws.title = "Live Detections"
        self._row   = 1
        self._count = 0
        self._last_save = time.time()
        self._setup()

    def _border(self):
        s = Side(border_style="thin", color="BBBBBB")
        return Border(left=s, right=s, top=s, bottom=s)

    def _setup(self):
        ws = self._ws

        # Title row
        ws.merge_cells("A1:I1")
        tc = ws["A1"]
        tc.value = (
            f"🚗  Live Plate Monitoring  |  "
            f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        )
        tc.font      = Font(bold=True, size=12, color="FFFFFF")
        tc.fill      = PatternFill("solid", fgColor="1F4E79")
        tc.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 26
        self._row = 2

        # Header row
        b = self._border()
        for col, hdr in enumerate(self.HEADERS, start=1):
            c = ws.cell(row=self._row, column=col, value=hdr)
            c.font      = Font(bold=True, color="FFFFFF", size=10)
            c.fill      = PatternFill("solid", fgColor="2E75B6")
            c.alignment = Alignment(horizontal="center",
                                    vertical="center", wrap_text=True)
            c.border = b
        ws.row_dimensions[self._row].height = 20
        self._row = 3

        # Column widths
        for i, w in enumerate([5, 16, 16, 14, 10, 20, 20, 12, 20], start=1):
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(i)
            ].width = w

        ws.freeze_panes = "A3"
        self._wb.save(self.path)

    def log(self, plate: TrackedPlate):
        with self._lock:
            self._count += 1
            ws = self._ws
            r  = self._row
            b  = self._border()

            duration = plate.last_seen - plate.first_seen

            row_fill = self.VALID_FILL if plate.confirmed else (
                self.ALT_FILL if r % 2 == 0 else None
            )

            first_str = datetime.fromtimestamp(plate.first_seen).strftime("%H:%M:%S")
            last_str  = datetime.fromtimestamp(plate.last_seen).strftime("%H:%M:%S")

            values = [
                self._count,
                plate.confirmed or "N/A",
                plate.state,
                f"{plate.conf_score:.1f}",
                plate.seen_count,
                first_str,
                last_str,
                f"{duration:.1f}",
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            ]

            for col, val in enumerate(values, start=1):
                cell = ws.cell(row=r, column=col, value=val)
                cell.border    = b
                cell.alignment = Alignment(horizontal="center",
                                           vertical="center")
                if row_fill:
                    cell.fill = row_fill
                if col == 2:
                    cell.font = Font(bold=True, size=11)

            ws.row_dimensions[r].height = 17
            self._row += 1

    def auto_save(self):
        with self._lock:
            try:
                self._wb.save(self.path)
                self._last_save = time.time()
            except Exception as e:
                print(f"  [EXCEL] Save error: {e}")

    def save(self):
        self.auto_save()

# ══════════════════════════════════════════════════════════════
# PLATE EXTRACTION (same logic as main.py)
# ══════════════════════════════════════════════════════════════

def force_lll_ddd_ll(chars: str) -> str:
    out = []
    for i, ch in enumerate(chars[:8]):
        if i < 3 or i >= 6:
            out.append(DIGIT_TO_LETTER.get(ch, ch) if ch.isdigit() else ch)
        else:
            out.append(LETTER_TO_DIGIT.get(ch, ch) if ch.isalpha() else ch)
    return "".join(out)

def clean_noise(text: str) -> str:
    t = text.upper()
    for w in NOISE_WORDS:
        t = t.replace(w, " ")
    return re.sub(r'\s+', ' ', t).strip()

def extract_plate_from_ocr(texts: list) -> Tuple[str, bool]:
    """
    Try to extract LLL-DDD-LL from OCR text list.
    Returns (plate_string, is_valid).
    """
    for raw in texts:
        cleaned = clean_noise(raw)
        chars   = re.sub(r'[^A-Z0-9]', '', cleaned.upper())

        if len(chars) < 6:
            continue

        # Sliding window
        for start in range(max(1, len(chars) - 10)):
            window = chars[start : start + 8]
            if len(window) < 8:
                window = window.ljust(8, '0')
            corrected = force_lll_ddd_ll(window)
            if STRICT_PLATE.match(corrected):
                plate = f"{corrected[:3]}-{corrected[3:6]}-{corrected[6:8]}"
                return plate, True

        # Last resort: first 8
        if len(chars) >= 8:
            forced = force_lll_ddd_ll(chars[:8])
            if STRICT_PLATE.match(forced):
                plate = f"{forced[:3]}-{forced[3:6]}-{forced[6:8]}"
                return plate, True

    return "", False

def extract_state_from_ocr(texts: list) -> str:
    combined = " ".join(texts).upper().replace(" ", "")
    for state, aliases in VALID_STATES.items():
        for alias in aliases:
            if alias in combined:
                return state.capitalize()
    return "Unknown"

# ══════════════════════════════════════════════════════════════
# PREPROCESSING
# ══════════════════════════════════════════════════════════════

def preprocess_for_ocr(img):
    """Multi-variant preprocessing. Returns list of preprocessed images."""
    results = []
    h, w = img.shape[:2]

    # Variant 1: CLAHE + sharpen
    up   = cv2.resize(img, (w*2, h*2), interpolation=cv2.INTER_CUBIC)
    gray = cv2.cvtColor(up, cv2.COLOR_BGR2GRAY)
    gray = cv2.createCLAHE(3.0, (8,8)).apply(gray)
    gray = cv2.GaussianBlur(gray, (3,3), 0)
    k    = np.array([[-1,-1,-1],[-1,9,-1],[-1,-1,-1]])
    v1   = cv2.filter2D(cv2.cvtColor(gray, cv2.COLOR_GRAY2BGR), -1, k)
    results.append(v1)

    # Variant 2: Otsu binarization
    gray2 = cv2.cvtColor(up, cv2.COLOR_BGR2GRAY)
    gray2 = cv2.createCLAHE(5.0, (8,8)).apply(gray2)
    _, b  = cv2.threshold(gray2, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    v2    = cv2.cvtColor(b, cv2.COLOR_GRAY2BGR)
    results.append(v2)

    # Variant 3: Inverted
    _, b3 = cv2.threshold(gray2, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
    v3    = cv2.cvtColor(b3, cv2.COLOR_GRAY2BGR)
    results.append(v3)

    return results

# ══════════════════════════════════════════════════════════════
# PLATE SIMILARITY
# ══════════════════════════════════════════════════════════════

def plate_similarity(a: str, b: str) -> float:
    a = re.sub(r'[^A-Z0-9]', '', a.upper())
    b = re.sub(r'[^A-Z0-9]', '', b.upper())
    if not a or not b:
        return 0.0
    matches = sum(1 for x, y in zip(a, b) if x == y)
    return matches / max(len(a), len(b))

# ══════════════════════════════════════════════════════════════
# TEMPORAL VOTER
# ══════════════════════════════════════════════════════════════

def vote_plate(readings: List[str], confidences: List[float],
               agreement: float = 0.5) -> Tuple[str, float]:
    """
    Vote across multiple readings of the same plate.
    Returns (best_plate, avg_confidence).
    """
    if not readings:
        return "", 0.0

    if len(readings) == 1:
        return readings[0], confidences[0] if confidences else 0.0

    # Normalise: strip dashes
    cleaned = [re.sub(r'[^A-Z0-9]', '', r.upper()) for r in readings]
    max_len = max(len(c) for c in cleaned)

    voted_chars = []
    for pos in range(min(max_len, 8)):
        char_weights = {}
        for i, chars in enumerate(cleaned):
            if pos < len(chars):
                ch   = chars[pos]
                conf = confidences[i] if i < len(confidences) else 50.0
                char_weights[ch] = char_weights.get(ch, 0) + conf

        if not char_weights:
            voted_chars.append('0' if 3 <= pos <= 5 else 'X')
            continue

        winner = max(char_weights, key=char_weights.get)
        voted_chars.append(winner)

    raw = "".join(voted_chars)
    if len(raw) >= 8:
        corrected = force_lll_ddd_ll(raw[:8])
        if STRICT_PLATE.match(corrected):
            plate = f"{corrected[:3]}-{corrected[3:6]}-{corrected[6:8]}"
            avg_conf = sum(confidences) / len(confidences)
            return plate, avg_conf

    # Fallback: most common reading
    winner = Counter(readings).most_common(1)[0][0]
    avg    = sum(confidences) / len(confidences) if confidences else 0
    return winner, avg

# ══════════════════════════════════════════════════════════════
# LIVE MONITOR CLASS
# ══════════════════════════════════════════════════════════════

VALID_STATES = {
    "ABUJA":["FCT","ABJ","ABUJA"],"LAGOS":["LAG","LGS","LAGOS"],
    "KANO":["KAN","KANO"],"OGUN":["OGN","OGUN"],"OYO":["OYO"],
    "RIVERS":["RIV","RIVERS"],"KADUNA":["KAD","KADUNA"],
    "ENUGU":["ENU","ENUGU"],"DELTA":["DEL","DELTA"],"EDO":["EDO"],
    "ANAMBRA":["ANM","ANAMBRA"],"IMO":["IMO"],
    "AKWAIBOM":["AKW","AKWAIBOM"],"CROSSRIVER":["CRS","CROSSRIVER"],
    "BORNO":["BOR","BORNO"],"NIGER":["NIG","NIGER"],
    "PLATEAU":["PLT","PLATEAU"],"KWARA":["KWR","KWARA"],
    "EKITI":["EKT","EKITI"],"OSUN":["OSN","OSUN"],
    "ONDO":["OND","ONDO"],"BAYELSA":["BAY","BAYELSA"],
    "ZAMFARA":["ZAM","ZAMFARA"],"KEBBI":["KEB","KEBBI"],
    "SOKOTO":["SOK","SOKOTO"],"YOBE":["YOB","YOBE"],
    "GOMBE":["GOM","GOMBE"],"NASARAWA":["NAS","NASARAWA"],
    "TARABA":["TAR","TARABA"],"JIGAWA":["JIG","JIGAWA"],
    "KOGI":["KOG","KOGI"],"BENUE":["BEN","BENUE"],
    "EBONYI":["EBO","EBONYI"],"ADAMAWA":["ADA","ADAMAWA"],
    "BAUCHI":["BAU","BAUCHI"],"KATSINA":["KAT","KATSINA"],
}

class LiveMonitor:
    """
    Main live monitoring engine.
    Captures frames → YOLO detect → OCR → Track → Vote → Log
    """

    def __init__(self, config: dict):
        self.config  = config
        self.running = False

        # ── Load models ──────────────────────────────────────
        print("\n[LIVE] Loading models...")
        yolo_path = os.path.join(BACKEND_DIR, config.get("yolo_model", "best.pt"))
        if not os.path.exists(yolo_path):
            raise FileNotFoundError(f"YOLO model not found: {yolo_path}")

        self.yolo = YOLO(yolo_path)
        print("[LIVE] ✅ YOLO loaded")

        try:
            self.ocr = PaddleOCR(lang="en", show_log=False)
            _ = self.ocr.ocr(np.ones((50,150,3), dtype=np.uint8)*255)
            print("[LIVE] ✅ PaddleOCR loaded")
        except Exception:
            try:
                self.ocr = PaddleOCR(use_angle_cls=False, lang="en", show_log=False)
                print("[LIVE] ✅ PaddleOCR loaded (legacy)")
            except Exception as e:
                print(f"[LIVE] ❌ PaddleOCR failed: {e}")
                self.ocr = None

        # ── State ────────────────────────────────────────────
        self.tracked: dict = {}   # track_id → TrackedPlate
        self.next_id       = 0
        self.frame_count   = 0
        self.detections    = 0
        self.start_time    = 0.0

        # ── Logger ───────────────────────────────────────────
        self.logger = LiveExcelLogger(config.get("excel_log", "live_log.xlsx"))

        # ── Stats ────────────────────────────────────────────
        self.confirmed_plates: List[str] = []

    def _run_ocr(self, img) -> Tuple[List[str], List[float]]:
        """Run OCR on image, return (texts, scores)."""
        if self.ocr is None:
            return [], []
        try:
            if hasattr(self.ocr, 'ocr'):
                res = self.ocr.ocr(img, cls=False)
            else:
                res = self.ocr.predict(img)

            texts, scores = [], []
            if not res:
                return texts, scores

            block = res[0] if isinstance(res, list) else res

            if isinstance(block, dict):
                texts  = block.get("rec_texts",  [])
                scores = block.get("rec_scores", [])
            elif isinstance(block, list):
                for line in block:
                    if not line or len(line) < 2:
                        continue
                    part = line[1]
                    if isinstance(part, (list, tuple)) and len(part) == 2:
                        t, s = part
                    else:
                        t, s = line[0], line[1]
                    if isinstance(t, str) and t.strip():
                        texts.append(t.strip())
                        scores.append(float(s))

            return texts, scores
        except Exception:
            return [], []

    def _process_crop(self, crop) -> Tuple[str, float, str]:
        """
        Process a cropped plate image.
        Returns (plate_number, confidence, state).
        """
        h = crop.shape[0]

        # Split regions
        top = crop[0 : int(h*0.25), :]
        mid = crop[int(h*0.35) : int(h*0.75), :]

        # Preprocess mid region variants
        variants    = preprocess_for_ocr(mid)
        all_texts   = []
        all_scores  = []

        for variant in variants:
            bordered = cv2.copyMakeBorder(
                variant, 15, 15, 15, 15,
                cv2.BORDER_CONSTANT, value=[255,255,255]
            )
            texts, scores = self._run_ocr(bordered)
            all_texts.extend(texts)
            all_scores.extend(scores)

        # Also try full crop
        full_variants = preprocess_for_ocr(crop)
        for fv in full_variants[:1]:
            bordered = cv2.copyMakeBorder(
                fv, 15, 15, 15, 15,
                cv2.BORDER_CONSTANT, value=[255,255,255]
            )
            texts, scores = self._run_ocr(bordered)
            all_texts.extend(texts)
            all_scores.extend(scores)

        # Extract plate
        plate, valid = extract_plate_from_ocr(all_texts)

        if not valid or not plate:
            return "Not Found", 0.0, "Unknown"

        # Confidence
        conf = sum(all_scores) / len(all_scores) * 100 if all_scores else 0

        # State
        top_texts, _ = self._run_ocr(
            cv2.copyMakeBorder(
                preprocess_for_ocr(top)[0] if top.size > 0
                else np.ones((50,100,3), dtype=np.uint8)*255,
                10,10,10,10,cv2.BORDER_CONSTANT,value=[255,255,255]
            )
        )
        state = extract_state_from_ocr(top_texts + all_texts)

        return plate, conf, state

    def _find_matching_track(self, plate: str) -> Optional[int]:
        """Find an existing track that matches this plate."""
        thresh = self.config.get("similarity_thresh", 0.6)
        for tid, track in self.tracked.items():
            if track.readings:
                best_reading = Counter(track.readings).most_common(1)[0][0]
                if plate_similarity(plate, best_reading) >= thresh:
                    return tid
        return None

    def _cleanup_tracks(self):
        """Remove expired tracks and log confirmed ones."""
        now     = time.time()
        timeout = self.config.get("track_timeout", 5.0)
        min_seen = self.config.get("min_seen_count", 2)
        expired = []

        for tid, track in self.tracked.items():
            if now - track.last_seen > timeout:
                # Plate left the frame — finalize
                if track.seen_count >= min_seen and not track.logged:
                    # Vote for best plate
                    voted, conf = vote_plate(
                        track.readings,
                        track.confidences,
                        self.config.get("vote_agreement", 0.5)
                    )
                    if voted and voted != "Not Found":
                        track.confirmed  = voted
                        track.conf_score = conf

                        # Enrich with state
                        try:
                            info = enrich_plate_data(voted)
                            track.state = info.get("plate_state", "Unknown")
                        except Exception:
                            track.state = get_state_from_plate(voted) or "Unknown"

                        # Log to Excel
                        track.logged = True
                        self.logger.log(track)
                        self.confirmed_plates.append(voted)

                        print(
                            f"\n  📋  CONFIRMED: {voted}  "
                            f"State={track.state}  "
                            f"Seen={track.seen_count}x  "
                            f"Conf={conf:.1f}%"
                        )

                expired.append(tid)

        for tid in expired:
            del self.tracked[tid]

    def _draw_overlay(self, frame, detections_info: list):
        """Draw bounding boxes and plate info on frame."""
        cfg = self.config

        for info in detections_info:
            x1, y1, x2, y2 = info["bbox"]
            plate = info.get("plate", "")
            conf  = info.get("conf", 0)
            state = info.get("state", "")
            seen  = info.get("seen", 0)

            # Box
            cv2.rectangle(
                frame, (x1, y1), (x2, y2),
                cfg["box_color"], cfg["box_thickness"]
            )

            # Text background
            label = f"{plate} | {state} | {conf:.0f}% | {seen}x"
            (tw, th), _ = cv2.getTextSize(
                label, cv2.FONT_HERSHEY_SIMPLEX, 0.6, 1
            )
            cv2.rectangle(
                frame,
                (x1, y1 - th - 10),
                (x1 + tw + 10, y1),
                cfg["text_bg_color"], -1
            )

            # Text
            cv2.putText(
                frame, label,
                (x1 + 5, y1 - 5),
                cv2.FONT_HERSHEY_SIMPLEX, 0.6,
                cfg["text_color"], 1, cv2.LINE_AA
            )

        # ── Status bar ───────────────────────────────────────
        elapsed = time.time() - self.start_time if self.start_time else 0
        fps     = self.frame_count / elapsed if elapsed > 0 else 0

        status = (
            f"Frames: {self.frame_count}  |  "
            f"FPS: {fps:.1f}  |  "
            f"Tracking: {len(self.tracked)}  |  "
            f"Confirmed: {len(self.confirmed_plates)}  |  "
            f"Press 'q' to quit"
        )
        cv2.rectangle(frame, (0, 0), (frame.shape[1], 30), (0, 0, 0), -1)
        cv2.putText(
            frame, status, (10, 20),
            cv2.FONT_HERSHEY_SIMPLEX, 0.5,
            (0, 255, 0), 1, cv2.LINE_AA
        )

        return frame

    def run(self):
        """Main loop: capture → detect → OCR → track → display → log."""

        source = self.config["source"]

        print(f"\n{'═'*60}")
        print(f"  🎥  LIVE PLATE MONITORING")
        print(f"  📹  Source: {source}")
        print(f"  📊  Log: {self.config['excel_log']}")
        print(f"  ⌨️   Press 'q' to stop")
        print(f"{'═'*60}\n")

        cap = cv2.VideoCapture(source)

        if not cap.isOpened():
            print(f"  ❌  Cannot open source: {source}")
            return

        # Set resolution if webcam
        if isinstance(source, int):
            cap.set(cv2.CAP_PROP_FRAME_WIDTH,  self.config["window_width"])
            cap.set(cv2.CAP_PROP_FRAME_HEIGHT, self.config["window_height"])

        self.running    = True
        self.start_time = time.time()
        process_every   = self.config.get("process_every_n", 3)
        last_save       = time.time()

        print("  ✅  Camera opened. Starting detection...\n")

        try:
            while self.running:
                ret, frame = cap.read()
                if not ret:
                    if isinstance(source, str) and os.path.isfile(source):
                        # Video file ended → loop
                        cap.set(cv2.CAP_PROP_POS_FRAMES, 0)
                        continue
                    print("  ❌  Lost connection to camera")
                    break

                self.frame_count += 1
                detections_info = []

                # Process every Nth frame
                if self.frame_count % process_every == 0:
                    try:
                        dets  = self.yolo(
                            frame,
                            imgsz=self.config["imgsz"],
                            verbose=False,
                            conf=self.config["min_yolo_conf"],
                        )
                        boxes = dets[0].boxes if dets else []
                    except Exception:
                        boxes = []

                    for box in boxes:
                        yolo_conf = float(box.conf[0])
                        x1, y1, x2, y2 = map(int, box.xyxy[0])

                        # Crop plate
                        h, w = frame.shape[:2]
                        pad  = self.config["crop_pad"]
                        cy1  = max(0, y1-pad)
                        cy2  = min(h, y2+pad)
                        cx1  = max(0, x1-pad)
                        cx2  = min(w, x2+pad)
                        crop = frame[cy1:cy2, cx1:cx2]

                        if crop.size == 0:
                            continue

                        ch, cw = crop.shape[:2]
                        if cw < 30 or ch < 15:
                            continue

                        # OCR
                        plate, conf, state = self._process_crop(crop)

                        if plate == "Not Found" or not plate:
                            continue

                        if conf < self.config["min_confidence"]:
                            continue

                        self.detections += 1

                        # ── Track ────────────────────────────
                        existing_tid = self._find_matching_track(plate)

                        if existing_tid is not None:
                            self.tracked[existing_tid].add_reading(
                                plate, conf, (x1, y1, x2, y2)
                            )
                            track = self.tracked[existing_tid]
                        else:
                            self.next_id += 1
                            track = TrackedPlate()
                            track.add_reading(plate, conf, (x1, y1, x2, y2))
                            track.state = state
                            self.tracked[self.next_id] = track

                        # Build display info
                        voted, _ = vote_plate(
                            track.readings[-10:],  # last 10 readings
                            track.confidences[-10:],
                        )
                        detections_info.append({
                            "bbox"  : (x1, y1, x2, y2),
                            "plate" : voted or plate,
                            "conf"  : conf,
                            "state" : track.state,
                            "seen"  : track.seen_count,
                        })

                # Cleanup expired tracks
                self._cleanup_tracks()

                # Auto-save Excel
                if time.time() - last_save > self.config.get("auto_save_interval", 30):
                    self.logger.auto_save()
                    last_save = time.time()

                # Draw overlay and display
                if self.config.get("show_window", True):
                    display = self._draw_overlay(frame.copy(), detections_info)

                    # Resize for display
                    dw = self.config["window_width"]
                    dh = self.config["window_height"]
                    display = cv2.resize(display, (dw, dh))

                    cv2.imshow("Live Plate Monitor", display)

                    key = cv2.waitKey(1) & 0xFF
                    if key == ord('q') or key == 27:  # q or ESC
                        print("\n  ⏹️  Stopping...")
                        self.running = False

        except KeyboardInterrupt:
            print("\n  ⏹️  Interrupted...")

        finally:
            cap.release()
            if self.config.get("show_window", True):
                cv2.destroyAllWindows()

            # Final save
            self.logger.save()

            # Print summary
            elapsed = time.time() - self.start_time if self.start_time else 0
            fps     = self.frame_count / elapsed if elapsed > 0 else 0

            print(f"\n{'═'*60}")
            print(f"  📊  SESSION SUMMARY")
            print(f"{'═'*60}")
            print(f"  Total Frames   : {self.frame_count}")
            print(f"  Duration       : {elapsed:.1f}s")
            print(f"  Avg FPS        : {fps:.1f}")
            print(f"  Detections     : {self.detections}")
            print(f"  Confirmed      : {len(self.confirmed_plates)}")

            if self.confirmed_plates:
                print(f"\n  🚗  Confirmed Plates:")
                for i, p in enumerate(set(self.confirmed_plates), 1):
                    count = self.confirmed_plates.count(p)
                    state = get_state_from_plate(p) or "Unknown"
                    print(f"      [{i}] {p}  "
                          f"(State={state.capitalize()}, seen {count}x)")

            print(f"\n  📊  Excel log → {self.config['excel_log']}")
            print(f"{'═'*60}\n")

# ══════════════════════════════════════════════════════════════
# ENTRY POINT
# ══════════════════════════════════════════════════════════════

if __name__ == "__main__":

    print("\n" + "═"*60)
    print("  🎥  NIGERIAN PLATE RECOGNITION — LIVE MONITOR")
    print("  📋  Format: LLL-DDD-LL  |  Logs to live_log.xlsx")
    print("═"*60)

    monitor = LiveMonitor(CONFIG)
    monitor.run()