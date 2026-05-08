# test_video.py
# Nigerian plate video analyser — strict LLL-DDD-LL format

import requests
import os
import re
import time
import json
import signal
import threading
from datetime import datetime
from typing import Optional, List, Tuple
from collections import Counter
from dataclasses import dataclass, field

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)

from statelga import (
    enrich_plate_data,
    validate_plate_format,
    get_state_from_plate,
    format_and_validate,
    correct_plate_characters,
)

# ══════════════════════════════════════════════════════════════
# CONFIGURATION
# ══════════════════════════════════════════════════════════════

CONFIG = {
    "base_url"            : "https://humble-palm-tree-gvj6vgvgg4j3pv6p-5000.app.github.dev/",
    "video_path"          : "/workspaces/AVLPR-DL/test_video/videoStreamTest2.mp4",
    "excel_log"           : "video_log.xlsx",
    "debug_log_file"      : "video_plate_debug.txt",
    "max_retries"         : 3,
    "retry_delay"         : 5,
    "upload_timeout"      : 600,
    "backend_timeout"     : 10,
    "min_confidence"      : 40,      # lower to catch more frames
    "min_seen_count"      : 1,
    "skip_frames"         : 10,
    "max_frames"          : 100,
    "strict_format"       : True,    # ONLY accept LLL-DDD-LL
    "similarity_threshold": 0.55,
    "vote_agreement"      : 0.50,
}

# ══════════════════════════════════════════════════════════════
# EXACT NIGERIAN PLATE FORMAT: LLL-DDD-LL
# ══════════════════════════════════════════════════════════════

# The ONLY valid format: 3 letters, 3 digits, 2 letters
# Examples: AFZ-169-EA, KJA-446-AC, FGB-778-JS

STRICT_PATTERN = re.compile(r'^[A-Z]{3}\d{3}[A-Z]{2}$')

# OCR confuses these characters — fix by position
# pos 0,1,2 → must be LETTERS  (fix digits → letters)
# pos 3,4,5 → must be DIGITS   (fix letters → digits)
# pos 6,7   → must be LETTERS  (fix digits → letters)

DIGIT_TO_LETTER = {
    '0': 'O', '1': 'I', '2': 'Z', '3': 'B',
    '4': 'A', '5': 'S', '6': 'G', '7': 'T',
    '8': 'B', '9': 'D',
}

LETTER_TO_DIGIT = {
    'O': '0', 'I': '1', 'L': '1', 'Z': '2',
    'B': '8', 'S': '5', 'G': '6', 'T': '7',
    'A': '4', 'E': '3', 'D': '0', 'Q': '0',
}

NOISE_WORDS = [
    "FEDERAL", "REPUBLIC", "NIGERIA", "CENTRE", "CENTER",
    "UNITY", "EXCELLENCE", "STATE", "GOVERNMENT", "OF",
    "LAGOS", "ABUJA", "FCT", "KANO", "RIVERS",
]

# ══════════════════════════════════════════════════════════════
# GRACEFUL SHUTDOWN
# ══════════════════════════════════════════════════════════════

_shutdown_requested = False

def _handle_signal(sig, frame):
    global _shutdown_requested
    print("\n\n⚠️  Interrupt received...")
    _shutdown_requested = True

signal.signal(signal.SIGINT,  _handle_signal)
signal.signal(signal.SIGTERM, _handle_signal)

# ══════════════════════════════════════════════════════════════
# DATACLASS
# ══════════════════════════════════════════════════════════════

@dataclass
class PlateReading:
    text       : str
    confidence : float
    frame_id   : int
    timestamp  : float
    source     : str = ""

# ══════════════════════════════════════════════════════════════
# SPINNER
# ══════════════════════════════════════════════════════════════

class Spinner:
    def __init__(self, msg: str = "Working"):
        self.msg = msg
        self.spinning = False
        self._thread = None

    def _spin(self):
        frames = ["|", "/", "-", "\\"]
        i = 0
        while self.spinning:
            print(f"\r  {frames[i % len(frames)]}  {self.msg}...",
                  end="", flush=True)
            time.sleep(0.1)
            i += 1

    def start(self):
        self.spinning = True
        self._thread = threading.Thread(target=self._spin, daemon=True)
        self._thread.start()

    def stop(self, msg: str = ""):
        self.spinning = False
        if self._thread:
            self._thread.join()
        print(f"\r  ✅  {msg or self.msg + ' done'}" + " " * 30)

# ══════════════════════════════════════════════════════════════
# DEBUG LOGGER
# ══════════════════════════════════════════════════════════════

class DebugLogger:
    def __init__(self, path: str):
        self.path  = path
        self._lock = threading.Lock()
        with open(self.path, "w", encoding="utf-8") as f:
            f.write(f"DEBUG LOG — {datetime.now()}\n{'='*60}\n\n")

    def _write(self, line: str):
        with self._lock:
            with open(self.path, "a", encoding="utf-8") as f:
                f.write(line + "\n")

    def log_raw(self, idx: int, plate: dict, note: str = ""):
        self._write(
            f"[RAW #{idx}] "
            f"plate='{plate.get('plate_number','')}' "
            f"conf={plate.get('confidence', 0):.1f}% "
            f"frame='{plate.get('filename','')}' "
            f"note={note}"
        )

    def log_section(self, title: str):
        self._write(f"\n{'='*60}\n{title}\n{'='*60}")

    def log_correction(self, original: str, corrected: str, reason: str = ""):
        self._write(
            f"[CORRECT] '{original}' → '{corrected}' {reason}"
        )

    def log_rejection(self, plate: str, reason: str):
        self._write(f"[REJECT] '{plate}' → {reason}")

    def log_vote(self, group_id: int, candidates: list,
                 winner: str, details: str = ""):
        self._write(
            f"[VOTE] group={group_id} "
            f"n={len(candidates)} "
            f"winner='{winner}' {details}"
        )

# ══════════════════════════════════════════════════════════════
# EXCEL LOGGER  (no summary block)
# ══════════════════════════════════════════════════════════════

class ExcelLogger:
    """
    Logs detected plates to video_log.xlsx.
    Columns: #, Plate Number, State, Confidence, Seen,
             First Seen, Last Seen, Format, Timestamp
    No summary block — clean data only.
    """

    HEADERS = [
        "#", "Plate Number", "State",
        "Confidence (%)", "Seen",
        "First Seen", "Last Seen",
        "Format", "Timestamp",
    ]

    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    VALID_FILL  = PatternFill("solid", fgColor="E2EFDA")
    ALT_FILL    = PatternFill("solid", fgColor="F5F5F5")

    def __init__(self, path: str, video_name: str = ""):
        self.path       = path
        self.video_name = video_name
        self._lock      = threading.Lock()
        self._wb        = openpyxl.Workbook()
        self._ws        = self._wb.active
        self._ws.title  = "Detected Plates"
        self._row       = 1
        self._setup_sheet()

    def _border(self):
        s = Side(border_style="thin", color="BBBBBB")
        return Border(left=s, right=s, top=s, bottom=s)

    def _setup_sheet(self):
        ws = self._ws

        # Title row
        ws.merge_cells("A1:I1")
        tc = ws["A1"]
        tc.value = (
            f"🚗  Nigerian Plate Recognition  |  "
            f"{self.video_name}  |  "
            f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        )
        tc.font      = Font(bold=True, size=12, color="FFFFFF")
        tc.fill      = PatternFill("solid", fgColor="1F4E79")
        tc.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 26
        self._row = 2

        # Header row
        b = self._border()
        for col, hdr in enumerate(self.HEADERS, start=1):
            c = ws.cell(row=self._row, column=col, value=hdr)
            c.font      = Font(bold=True, color="FFFFFF", size=10)
            c.fill      = PatternFill("solid", fgColor="2E75B6")
            c.alignment = Alignment(horizontal="center",
                                    vertical="center", wrap_text=True)
            c.border    = b
        ws.row_dimensions[self._row].height = 20
        self._row = 3

        # Column widths
        for i, w in enumerate([5, 16, 16, 14, 8, 22, 22, 14, 20], start=1):
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(i)
            ].width = w

        ws.freeze_panes = "A3"

    def log(self, idx: int, entry: dict):
        with self._lock:
            self._write_row(idx, entry)
            self._wb.save(self.path)

    def _write_row(self, idx: int, e: dict):
        ws = self._ws
        r  = self._row
        b  = self._border()

        is_valid = e.get("format_valid", False)
        row_fill = self.VALID_FILL if is_valid else (
            self.ALT_FILL if r % 2 == 0 else None
        )

        values = [
            idx,
            e.get("plate_number",  "N/A"),
            e.get("plate_state",   "Unknown"),
            f"{e.get('confidence', 0):.1f}",
            e.get("seen_count",    1),
            e.get("first_seen",    "N/A"),
            e.get("last_seen",     "N/A"),
            e.get("plate_format",  "N/A"),
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        ]

        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.border    = b
            cell.alignment = Alignment(horizontal="center",
                                       vertical="center")
            if row_fill:
                cell.fill = row_fill
            if col == 2:
                cell.font = Font(bold=True, size=11)

        ws.row_dimensions[r].height = 17
        self._row += 1

    def save(self):
        self._wb.save(self.path)

# ══════════════════════════════════════════════════════════════
# CORE: POSITION-AWARE CORRECTION  (LLL-DDD-LL)
# ══════════════════════════════════════════════════════════════

def force_lll_ddd_ll(chars: str) -> str:
    """
    Force an 8-char alphanumeric string into LLL-DDD-LL.
    - pos 0,1,2 → must be LETTERS
    - pos 3,4,5 → must be DIGITS
    - pos 6,7   → must be LETTERS
    """
    out = []
    for i, ch in enumerate(chars[:8]):
        if i < 3 or i >= 6:
            # Letter zone: fix digits
            out.append(DIGIT_TO_LETTER.get(ch, ch) if ch.isdigit() else ch)
        else:
            # Digit zone: fix letters
            out.append(LETTER_TO_DIGIT.get(ch, ch) if ch.isalpha() else ch)
    return "".join(out)

def clean_noise(text: str) -> str:
    """Strip plate-frame slogans before processing."""
    t = text.upper()
    for w in NOISE_WORDS:
        t = t.replace(w, " ")
    return re.sub(r'\s+', ' ', t).strip()

def extract_lll_ddd_ll(raw: str, debug_logger: DebugLogger,
                      label: str = "") -> Optional[str]:
    """
    STRICT extractor for Nigerian plates (LLL-DDD-LL ONLY)

    Stronger than before:
    ✔ scans ALL possible windows
    ✔ prioritizes valid patterns
    ✔ removes noise aggressively
    ✔ NEVER returns RAW garbage
    """

    # 1. Clean noise (VERY important)
    text = clean_noise(raw)
    chars = re.sub(r'[^A-Z0-9]', '', text.upper())

    debug_logger._write(
        f"[EXTRACT] raw='{raw}' → cleaned='{chars}'"
    )

    if len(chars) < 6:
        debug_logger.log_rejection(raw, "Too short")
        return None

    best_match = None

    # 2. Scan ALL possible windows (FIXED RANGE)
    for start in range(0, len(chars)):
        window = chars[start:start + 8]

        if len(window) < 6:
            continue

        # Pad to 8 if needed
        window = window.ljust(8, '0')

        corrected = force_lll_ddd_ll(window)

        debug_logger.log_correction(
            window, corrected, f"(start={start})"
        )

        # STRICT MATCH ONLY
        if STRICT_PATTERN.match(corrected):
            best_match = corrected
            break

    # 3. If found → return formatted
    if best_match:
        plate = f"{best_match[:3]}-{best_match[3:6]}-{best_match[6:8]}"
        debug_logger._write(f"[EXTRACT] ✅ MATCH='{plate}'")
        return plate

    # ❌ NO FALLBACK TO RAW ANYMORE
    debug_logger.log_rejection(raw, "No valid LLL-DDD-LL found")
    return None

# ══════════════════════════════════════════════════════════════
# SIMILARITY & GROUPING
# ══════════════════════════════════════════════════════════════

def plate_similarity(a: str, b: str) -> float:
    """
    Compare two LLL-DDD-LL plates character by character.
    Returns 0.0–1.0.
    """
    a = re.sub(r'[^A-Z0-9]', '', a.upper())
    b = re.sub(r'[^A-Z0-9]', '', b.upper())
    if not a or not b:
        return 0.0
    length  = max(len(a), len(b))
    matches = sum(1 for x, y in zip(a, b) if x == y)
    return matches / length

def group_similar_plates(plates: list, threshold: float = 0.55) -> list:
    """Group OCR readings of the same physical plate."""
    groups = []
    for plate in plates:
        placed = False
        for group in groups:
            if plate_similarity(
                plate["corrected"], group[0]["corrected"]
            ) >= threshold:
                group.append(plate)
                placed = True
                break
        if not placed:
            groups.append([plate])
    return groups

def split_plate_regions(img):
    """
    VERY STRICT split based on Nigerian plate layout:

    ┌──────────────┐
    │   STATE      │  ← ignore mostly
    │──────────────│
    │  ABC-123-DE  │  ← THIS is what we want
    │──────────────│
    │  small text  │  ← ignore
    └──────────────┘
    """

    h = img.shape[0]

    # Narrow band where plate numbers ACTUALLY are
    plate_band = img[int(h * 0.40): int(h * 0.75), :]

    # Smaller state region
    state_band = img[0: int(h * 0.25), :]

    return state_band, plate_band

def filter_plate_candidates(texts):
    """
    Keep ONLY strings that look like plate fragments
    """
    candidates = []

    for t in texts:
        t = re.sub(r'[^A-Z0-9]', '', t.upper())

        # Must contain BOTH letters and digits
        if (re.search(r'[A-Z]', t) and re.search(r'\d', t)):
            if 6 <= len(t) <= 10:
                candidates.append(t)

    return candidates      

# ══════════════════════════════════════════════════════════════
# TEMPORAL VOTER  (character-level, position-aware)
# ══════════════════════════════════════════════════════════════

class TemporalVoter:
    """
    Given multiple readings of the same plate across frames,
    vote character-by-character to find the correct plate.
    Then apply force_lll_ddd_ll to ensure valid format.
    """

    def __init__(self, config: dict):
        self.vote_agreement = config.get("vote_agreement", 0.50)
        self.min_confidence = config.get("min_confidence", 40)

    def aggregate(self, readings: List[PlateReading],
                  debug_logger: DebugLogger) -> Tuple[Optional[str], float]:

        if not readings:
            return None, 0.0

        # Filter by confidence
        valid = [r for r in readings if r.confidence >= self.min_confidence]
        if not valid:
            valid = readings

        # Normalise: strip dashes, uppercase, exactly 8 chars
        normalised = []
        for r in valid:
            c = re.sub(r'[^A-Z0-9]', '', r.text.upper())
            if len(c) >= 8:
                normalised.append((c[:8], r.confidence))
            elif len(c) >= 6:
                normalised.append((c.ljust(8, '0'), r.confidence))

        if not normalised:
            # Fallback: just return highest confidence reading
            best = max(valid, key=lambda r: r.confidence)
            return best.text, best.confidence

        # Character-level voting (weighted by confidence)
        voted_chars = []
        for pos in range(8):
            char_votes: dict = {}
            for chars, conf in normalised:
                ch = chars[pos] if pos < len(chars) else '_'
                if ch == '_':
                    continue
                char_votes[ch] = char_votes.get(ch, 0) + conf

            if not char_votes:
                voted_chars.append('0' if 3 <= pos <= 5 else 'X')
                continue

            winner = max(char_votes, key=char_votes.get)
            total  = sum(char_votes.values())
            agreement = char_votes[winner] / total if total > 0 else 0

            debug_logger._write(
                f"  [VOTE pos={pos}] votes={char_votes} "
                f"winner='{winner}' agree={agreement:.0%}"
            )

            voted_chars.append(winner)

        raw_voted = "".join(voted_chars)
        corrected = force_lll_ddd_ll(raw_voted)

        debug_logger._write(
            f"  [VOTE result] raw='{raw_voted}' corrected='{corrected}'"
        )

        # Confidence = average of all valid readings
        avg_conf = sum(conf for _, conf in normalised) / len(normalised)
        return corrected, avg_conf

# ══════════════════════════════════════════════════════════════
# VOTE BEST PLATE
# ══════════════════════════════════════════════════════════════

def vote_best_plate(
    group       : list,
    debug_logger: DebugLogger,
    voter       : TemporalVoter,
    group_id    : int = 0,
) -> Optional[dict]:

    readings = [
        PlateReading(
            text       = e["corrected"],
            confidence = e.get("confidence", 0),
            frame_id   = 0,
            timestamp  = time.time(),
            source     = e.get("source", ""),
        )
        for e in group
    ]

    debug_logger.log_section(
        f"GROUP {group_id}  n={len(group)}  "
        f"plates={[e['corrected'] for e in group]}"
    )

    voted_raw, vote_conf = voter.aggregate(readings, debug_logger)

    if not voted_raw:
        debug_logger.log_rejection(
            str([e["corrected"] for e in group]),
            "Voting returned None"
        )
        return None

    # Apply strict format
    chars     = re.sub(r'[^A-Z0-9]', '', voted_raw.upper())
    corrected = force_lll_ddd_ll(chars[:8]) if len(chars) >= 8 else chars
    is_valid  = STRICT_PATTERN.match(corrected) is not None

    if is_valid:
        winner = f"{corrected[:3]}-{corrected[3:6]}-{corrected[6:8]}"
    else:
        # Try each candidate individually
        for entry in sorted(group,
                            key=lambda e: e.get("confidence", 0),
                            reverse=True):
            cand = re.sub(r'[^A-Z0-9]', '', entry["corrected"].upper())
            forced = force_lll_ddd_ll(cand[:8]) if len(cand) >= 8 else cand
            if STRICT_PATTERN.match(forced):
                winner = f"{forced[:3]}-{forced[3:6]}-{forced[6:8]}"
                vote_conf = entry.get("confidence", vote_conf)
                debug_logger._write(
                    f"[VOTE] Used best candidate: '{winner}'"
                )
                is_valid = True
                break
        else:
            # No valid candidate — skip this group
            debug_logger.log_rejection(
                corrected, "Could not force to LLL-DDD-LL"
            )
            return None

    best_entry = next(
        (e for e in group if e["corrected"].replace("-", "") == corrected),
        group[0]
    )

    debug_logger.log_vote(
        group_id,
        [e["corrected"] for e in group],
        winner,
        f"conf={vote_conf:.1f}%"
    )

    return {
        "plate_number"  : winner,
        "confidence"    : round(vote_conf, 2),
        "detected_state": best_entry.get("state", "Unknown"),
        "source"        : best_entry.get("source", "N/A"),
        "seen_count"    : len(group),
        "first_seen"    : group[0].get("source", "N/A"),
        "last_seen"     : group[-1].get("source", "N/A"),
    }

# ══════════════════════════════════════════════════════════════
# PROCESS RESULTS
# ══════════════════════════════════════════════════════════════

def process_results(
    raw_plates  : list,
    debug_logger: DebugLogger,
    config      : dict,
) -> Tuple[list, dict]:

    min_conf   = config.get("min_confidence",       40)
    sim_thresh = config.get("similarity_threshold", 0.55)

    stats = {
        "skipped_total"  : 0,
        "accepted"       : 0,
        "rejected_vote"  : 0,
    }

    voter = TemporalVoter(config)

    # ── Dump raw ──────────────────────────────────────────────
    debug_logger.log_section("RAW PLATES FROM SERVER")
    for i, p in enumerate(raw_plates):
        debug_logger.log_raw(i, p)

    # ── Step 1: Extract & correct every plate ────────────────
    corrected_plates = []

    debug_logger.log_section("EXTRACTION")

    for plate in raw_plates:
        if _shutdown_requested:
            break

        plate_number = clean_noise(
            str(plate.get("plate_number", "") or "")
        ).upper()
        confidence   = float(plate.get("confidence",   0) or 0)
        state        = str(plate.get("state_of_origin", "Unknown") or "Unknown")
        source       = str(plate.get("filename",       "N/A") or "N/A")

        if not plate_number or plate_number in ("NOT FOUND", "N/A", "NONE", ""):
            stats["skipped_total"] += 1
            continue

        if confidence < min_conf:
            debug_logger.log_rejection(
                plate_number, f"conf={confidence:.1f}% < {min_conf}%"
            )
            stats["skipped_total"] += 1
            continue

        # Extract strict LLL-DDD-LL
        extracted = extract_lll_ddd_ll(
            plate_number, debug_logger, label=source
        )

        if not extracted:
            # Try again with raw (no noise removal)
            extracted = extract_lll_ddd_ll(
                plate.get("plate_number", ""),
                debug_logger,
                label=source + "_retry"
            )

        if not extracted:
            stats["skipped_total"] += 1
            continue

        corrected_plates.append({
            "corrected" : extracted,
            "raw"       : plate_number,
            "confidence": confidence,
            "state"     : state,
            "source"    : source,
        })
        stats["accepted"] += 1

    if not corrected_plates:
        debug_logger.log_section("NO PLATES PASSED EXTRACTION")
        print(f"\n  ⚠️  0 plates passed extraction")
        print(f"  💡  Check debug log for details: "
              f"{config.get('debug_log_file')}")
        return [], stats

    debug_logger.log_section(
        f"{len(corrected_plates)} PLATES PASSED EXTRACTION"
    )

    # ── Step 2: Group similar plates ─────────────────────────
    groups = group_similar_plates(corrected_plates, threshold=sim_thresh)
    debug_logger.log_section(f"{len(groups)} GROUPS AFTER SIMILARITY")

    # ── Step 3: Temporal vote per group ──────────────────────
    final_plates = []
    for gid, group in enumerate(groups):
        if _shutdown_requested:
            break
        winner = vote_best_plate(group, debug_logger, voter, gid)
        if winner:
            final_plates.append(winner)
        else:
            stats["rejected_vote"] += 1

    # ── Step 4: Enrich with state / LGA ──────────────────────
    enriched = []
    for data in final_plates:
        pn = data["plate_number"]
        try:
            info = enrich_plate_data(
                pn, detected_state=data["detected_state"]
            )
        except Exception as e:
            info = {
                "plate_number"  : pn,
                "plate_state"   : "Unknown",
                "detected_state": data.get("detected_state", "Unknown"),
                "state_match"   : False,
                "format_valid"  : True,
                "format_message": f"Enrich error: {e}",
                "plate_format"  : "LLL-DDD-LL",
                "lga_count"     : 0,
            }

        info.update({
            "confidence": data.get("confidence", 0),
            "source"    : data.get("source",     "N/A"),
            "seen_count": data.get("seen_count",  1),
            "first_seen": data.get("first_seen",  "N/A"),
            "last_seen" : data.get("last_seen",   "N/A"),
        })
        enriched.append(info)

    # Sort by confidence descending
    enriched.sort(key=lambda x: x.get("confidence", 0), reverse=True)
    return enriched, stats

# ══════════════════════════════════════════════════════════════
# BACKEND CHECK
# ══════════════════════════════════════════════════════════════

def check_backend(base_url: str, timeout: int = 10) -> bool:
    try:
        r = requests.get(
            f"{base_url.rstrip('/')}/api/test", timeout=timeout
        )
        if r.status_code == 200:
            d = r.json()
            print(
                f"  ✅  Backend healthy  "
                f"(YOLO={'✅' if d.get('yolo') else '❌'}  "
                f"OCR={'✅' if d.get('ocr') else '❌'})"
            )
            return True
        print(f"  ⚠️  Backend HTTP {r.status_code}")
        return False
    except requests.exceptions.ConnectionError:
        print("  ❌  Backend unreachable — run: python main.py")
        return False
    except Exception as e:
        print(f"  ❌  Backend error: {e}")
        return False

# ══════════════════════════════════════════════════════════════
# VIDEO UPLOAD
# ══════════════════════════════════════════════════════════════

def send_video(
    video_path  : str,
    base_url    : str,
    max_retries : int = 3,
    retry_delay : int = 5,
    timeout     : int = 600,
    skip_frames : int = 10,
    max_frames  : int = 100,
) -> Optional[dict]:

    url = f"{base_url.rstrip('/')}/api/process-video"

    for attempt in range(1, max_retries + 1):
        if _shutdown_requested:
            return None

        print(f"\n  📤  Upload attempt {attempt}/{max_retries}...")

        try:
            with open(video_path, "rb") as f:
                sp = Spinner("Uploading & processing")
                sp.start()
                resp = requests.post(
                    url,
                    files={"file": (os.path.basename(video_path),
                                    f, "video/mp4")},
                    data={"skip_frames": skip_frames,
                          "max_frames" : max_frames},
                    timeout=timeout,
                )
                sp.stop("Upload complete")

        except requests.exceptions.Timeout:
            print(f"  ❌  Attempt {attempt} timed out")
            if attempt < max_retries: time.sleep(retry_delay)
            continue
        except requests.exceptions.ConnectionError as e:
            print(f"  ❌  Connection lost: {e}")
            if attempt < max_retries: time.sleep(retry_delay)
            continue
        except FileNotFoundError:
            print(f"  ❌  File not found: {video_path}")
            return None
        except Exception as e:
            print(f"  ❌  Upload error: {e}")
            return None

        if resp.status_code != 200:
            print(f"  ❌  HTTP {resp.status_code}: {resp.text[:200]}")
            if attempt < max_retries: time.sleep(retry_delay)
            continue

        if not resp.text.strip():
            print("  ❌  Empty server response")
            if attempt < max_retries: time.sleep(retry_delay)
            continue

        try:
            return resp.json()
        except Exception:
            print("  ❌  JSON parse error")
            return None

    print(f"  ❌  All {max_retries} attempts failed.")
    return None

# ══════════════════════════════════════════════════════════════
# DISPLAY
# ══════════════════════════════════════════════════════════════

def print_header():
    print("\n" + "═"*60)
    print("  🚗  NIGERIAN LICENSE PLATE — VIDEO ANALYSER")
    print("  📋  Format: LLL-DDD-LL  |  Temporal Voting")
    print(f"  🕐  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("═"*60)

def print_plate_box(idx: int, p: dict):
    plate = p.get("plate_number", "N/A")
    state = p.get("plate_state",  "Unknown")
    conf  = f"{p.get('confidence', 0):.1f}%"
    seen  = f"{p.get('seen_count', 1)}x"
    first = str(p.get("first_seen", "N/A"))
    last  = str(p.get("last_seen",  "N/A"))
    if len(first) > 22: first = "..." + first[-19:]
    if len(last)  > 22: last  = "..." + last[-19:]

    print(f"\n  ┌──────────────────────────────────────────────┐")
    print(f"  │  🔎  Plate #{idx:<3}                               │")
    print(f"  ├──────────────────────────────────────────────┤")
    print(f"  │  Plate Number : {plate:<29}│")
    print(f"  │  State        : {state:<29}│")
    print(f"  │  Confidence   : {conf:<29}│")
    print(f"  │  Seen         : {seen:<29}│")
    print(f"  │  First Seen   : {first:<29}│")
    print(f"  │  Last Seen    : {last:<29}│")
    print(f"  └──────────────────────────────────────────────┘")

def print_summary(stats: dict):
    elapsed = stats.get("elapsed", 0)
    proc_f  = stats.get("processed_frames", 0)
    rate    = proc_f / elapsed if elapsed > 0 and isinstance(proc_f, int) else 0

    print("\n" + "═"*60)
    print("  📊  SUMMARY")
    print("═"*60)
    rows = [
        ("Total Frames",    "total_frames"),
        ("Processed",       "processed_frames"),
        ("Raw Plates",      "plates_found"),
        ("Unique Plates",   "unique_plates"),
        ("Valid Formats",   "valid_formats"),
        ("State Resolved",  "state_matches"),
        ("Skipped",         "skipped_total"),
        ("Vote Rejected",   "rejected_vote"),
    ]
    for label, key in rows:
        print(f"  {label:<20}: {stats.get(key, 'N/A')}")
    print(f"  {'Time':<20}: {elapsed:.1f}s  ({rate:.1f} fps)")
    print(f"  {'Excel':<20}: {stats.get('excel_log', 'N/A')}")
    print(f"  {'Debug':<20}: {stats.get('debug_file', 'N/A')}")
    print("═"*60)

# ══════════════════════════════════════════════════════════════
# MAIN RUN
# ══════════════════════════════════════════════════════════════

def run(video_path: str, base_url: str, config: dict):

    print_header()

    excel_file = config.get("excel_log",      "video_log.xlsx")
    debug_file = config.get("debug_log_file", "video_plate_debug.txt")
    video_name = os.path.basename(video_path)

    excel_logger = ExcelLogger(excel_file, video_name=video_name)
    debug_logger = DebugLogger(debug_file)

    if not os.path.exists(video_path):
        print(f"\n  ❌  Video not found: {video_path}")
        return

    size_mb = os.path.getsize(video_path) / (1024 * 1024)
    print(f"\n  🎬  Video : {video_name}")
    print(f"  📦  Size  : {size_mb:.2f} MB")

    print(f"\n  🔗  Backend : {base_url}")
    if not check_backend(base_url, config.get("backend_timeout", 10)):
        return

    print(f"\n  ⏳  Uploading video...")
    print(f"      skip={config.get('skip_frames',10)} frames  "
          f"max={config.get('max_frames',100)} frames")

    start  = time.time()
    result = send_video(
        video_path, base_url,
        max_retries = config.get("max_retries",    3),
        retry_delay = config.get("retry_delay",    5),
        timeout     = config.get("upload_timeout", 600),
        skip_frames = config.get("skip_frames",    10),
        max_frames  = config.get("max_frames",     100),
    )
    elapsed = time.time() - start

    if result is None:
        return

    raw_plates       = result.get("results",          [])
    total_frames     = result.get("total_frames",     "N/A")
    processed_frames = result.get("processed_frames", "N/A")
    plates_found     = result.get("plates_found",     0)

    print(f"\n  ✅  Server responded in {elapsed:.1f}s")
    print(f"  📊  Raw plates: {len(raw_plates)}")

    if raw_plates:
        print("\n  🔍  RAW PREVIEW (first 5):")
        for i, p in enumerate(raw_plates[:5]):
            print(
                f"      [{i}] "
                f"'{p.get('plate_number','?')}' "
                f"conf={p.get('confidence','?')}% "
                f"@ {p.get('filename','?')}"
            )
        if len(raw_plates) > 5:
            print(f"      ... +{len(raw_plates)-5} more → {debug_file}")
    else:
        print("  ⚠️  Server returned 0 plates.")
        return

    print(f"\n  🔄  Extracting LLL-DDD-LL plates + temporal voting...")

    enriched, skip_stats = process_results(
        raw_plates, debug_logger, config
    )

    print("\n" + "═"*60)
    print(f"  🚗  DETECTED PLATES  ({len(enriched)} unique)")
    print("═"*60)

    if not enriched:
        print("\n  ⚠️  No valid plates found.")
        print(f"  💡  Check debug log: {debug_file}")
    else:
        for idx, plate_info in enumerate(enriched, 1):
            if _shutdown_requested:
                break
            print_plate_box(idx, plate_info)
            excel_logger.log(idx, plate_info)

    excel_logger.save()
    print(f"\n  📊  Excel log → {excel_file}")
    print(f"  🐛  Debug log → {debug_file}")

    print_summary({
        "total_frames"    : total_frames,
        "processed_frames": processed_frames,
        "plates_found"    : plates_found,
        "unique_plates"   : len(enriched),
        "valid_formats"   : sum(1 for p in enriched if p.get("format_valid")),
        "state_matches"   : sum(1 for p in enriched if p.get("state_match")),
        "skipped_total"   : skip_stats.get("skipped_total",  0),
        "rejected_vote"   : skip_stats.get("rejected_vote",  0),
        "elapsed"         : elapsed,
        "excel_log"       : excel_file,
        "debug_file"      : debug_file,
    })

# ══════════════════════════════════════════════════════════════
# ENTRY POINT
# ══════════════════════════════════════════════════════════════

if __name__ == "__main__":
    run(
        video_path = CONFIG["video_path"],
        base_url   = CONFIG["base_url"],
        config     = CONFIG,
    )

